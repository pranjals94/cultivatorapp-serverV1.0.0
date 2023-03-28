import os
import sys
import openpyxl
from sqlalchemy import or_

sys.path.append("..")

from fastapi import Depends, APIRouter, Request, Response, HTTPException, UploadFile, File
from Modules import auth
from sqlalchemy.orm import Session
from models import model
from schemas import schema

router = APIRouter(
    prefix="/reception",
    tags=["reception"],
    responses={401: {"user": "Not Authorised"}},
)


@router.post("/assign-persons-to-cultivator")
async def get_guests(obj: schema.assign_cultivator_to_persons, db: Session = Depends(auth.get_db)):
    for item in obj.persons:
        create_person_model = db.get(model.Person, item)
        create_person_model.cultivator_id = obj.cultivator.id
        db.add(create_person_model)
        db.commit()
    return {"msg": f"Persons Assigned to CULTIVATOR {obj.cultivator.name}"}


@router.get("/getcultivators")
async def get_cultivators(request: Request, db: Session = Depends(auth.get_db)):
    # user = await auth.get_current_user(request)
    # if user is None:
    #     raise HTTPException(status_code=401, detail="Sorry you are Unauthorized !")
    tempCultivators = db.query(model.User).filter(model.User.role_id == 2).offset(0).limit(
        50).all()  # role id 2 is for cultivator
    # userDetails = db.get(model.User, user.get("id"))
    # print(cultivators[2].person.name)
    # if(userDetails.personRole.name!="ADMIN"):
    #     raise HTTPException(status_code=403, detail="Forbidden !")
    # personUser = db.query(model.Person).filter(model.Person.id == userDetails.person_id).first()
    cultivators: list = []
    for cultivator in tempCultivators:
        results = {"id": cultivator.person.id, "name": cultivator.person.name}
        cultivators.append(results)
    return {"cultivators": cultivators}


@router.get("/getguests")
async def get_guests(currentPage: int, pageSize: int, db: Session = Depends(auth.get_db)):
    offset = pageSize * (currentPage - 1)
    totalGuests = db.query(model.Person).filter(model.Person.cultivator_id == None).count()
    tempGuests = db.query(model.Person).filter(model.Person.cultivator_id == None).order_by(
        model.Person.id.desc()).offset(offset).limit(
        pageSize).all()
    guests: list = []
    for guest in tempGuests:
        results = {"id": guest.id, "name": guest.name, "phone_no": guest.phone_no, "email": guest.email,
                   "gender": guest.gender}
        guests.append(results)
    return {"guests": guests, "totalGuests": totalGuests}


@router.post("/importfromexcel")
async def import_from_excel(file: UploadFile = File(...), db: Session = Depends(auth.get_db)):
    # path = "C:\\Users\\dcdrns\\Desktop\\Books.xlsx" #syntax for path in python
    # filename, file_extension = os.path.splitext(file.filename) #split the file name and extension
    # file.filename = f"tempXLfile{file_extension}"  # rename the file

    file.filename = f"tempXLfile.xlsx"  # rename the file
    contents = await file.read()
    with open(f"{file.filename}", "wb") as f:
        f.write(contents)
    try:
        wb_obj = openpyxl.load_workbook("tempXLfile.xlsx")
    except:
        return {"msg": "Not a valid .xlsx file !"}

    check_fields = [
        "first_name",
        "middle_name",
        "last_name",
        "mobile_no",
        "check_in_date_time,"
        "check_out_date_time,"
        "accommodation"
    ]
    sheet_obj = wb_obj.active  # This is set to 0 by default. Unless you modify its value, you will always get the
    # first worksheet by using this method.

    for i in range(3, sheet_obj.max_row + 1):  # row
        if sheet_obj.cell(row=i, column=4).data_type != 'n' \
                or sheet_obj.cell(row=i, column=4).value is None \
                or sheet_obj.cell(row=i,
                                  column=1).value is None:  # name cannot be null phone_no cant' be null or non numeric
            return {
                "msg": f"Emty or Invalid Name and phone no. in Row: {i}. \n max_row: {sheet_obj.max_row} \n max_column: {sheet_obj.max_column}"}

    for i in range(2, sheet_obj.max_row + 1):  # row

        if i == 2:
            for j in range(1, sheet_obj.max_column + 1):  # itrate each column
                # cell_obj = sheet_obj['A2']
                cell_obj = sheet_obj.cell(row=i, column=j)
                if cell_obj.value != check_fields[j - 1] or sheet_obj.max_column > len(check_fields):
                    return {"msg": ".xlsx file sample not correct."}
        else:
            person = model.Person()
            person.name = sheet_obj.cell(row=i, column=1).value
            person.first_name = sheet_obj.cell(row=i, column=2).value
            person.last_name = sheet_obj.cell(row=i, column=3).value
            person.phone_no = sheet_obj.cell(row=i, column=4).value
            person.email = sheet_obj.cell(row=i, column=5).value
            person.city = sheet_obj.cell(row=i, column=6).value
            person.gender = sheet_obj.cell(row=i, column=7).value
            # person.dob = '1997-11-11 13:23:44'
            if sheet_obj.cell(row=i, column=8).is_date:
                person.dob = sheet_obj.cell(row=i, column=8).value

            db_person = db.query(model.Person).filter(
                model.Person.name == person.name).filter(model.Person.phone_no == person.phone_no).first()
            if db_person:
                return {"msg": f"Person with same name and phone no. Already Exist. Row: {i}."}

            db.add(person)
            db.commit()
            db.refresh(person)
    wb_obj.close()  # not necessary
    return {"msg": "Excel sheet uploaded !"}


@router.get("/search")
async def search(currentPage: int = 1, pageSize: int = 10, search_input: str = '', db: Session = Depends(auth.get_db)):
    searchData = f'%{search_input}%'
    if searchData == '%%':
        # return {"msg": "Empty search value."}
        raise HTTPException(status_code=404, detail="Empty Search value!")
    offset = pageSize * (currentPage - 1)
    persons = db.query(model.Person).filter(model.Person.cultivator_id == None).filter(
        or_(model.Person.name.ilike(searchData),
            model.Person.phone_no.ilike(
                searchData))).offset(offset).limit(
        pageSize).all()  # ilike gurantees case insensitive
    totalGuests = db.query(model.Person).filter(or_(model.Person.name.ilike(searchData),
                                                    model.Person.phone_no.ilike(
                                                        searchData))).count()
    return {"persons": persons, "totalGuests": totalGuests}
